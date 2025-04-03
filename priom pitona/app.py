from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import make_pipeline

import requests
from flask import Flask, request, jsonify

import openpyxl
from openpyxl import load_workbook

import pymorphy2

morph = pymorphy2.MorphAnalyzer()

app = Flask(__name__)

TOKEN = '7565599070:AAEwOj102OQanJq8liXBK1qZCrkKjcKld6w'

ADMIN_CHAT_ID = '1070122283'

user_data = {}

def load_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    requests = []
    categories = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        requests.append(row[0])  
        categories.append(row[1]) 

    return requests, categories

def preprocess_text(text):
    words = text.split()
    processed_words = [morph.parse(word)[0].normal_form for word in words]
    return ' '.join(processed_words)

file_path = 'table.xlsx'  
X, y = load_data_from_excel(file_path)
X = [preprocess_text(x) for x in X] 

model = make_pipeline(TfidfVectorizer(), MultinomialNB())
model.fit(X, y)


def predict_category(request):
    processed_request = preprocess_text(request)
    return model.predict([processed_request])[0]


@app.route('/predict', methods=['POST'])
def predict():
    data = request.get_json()
    user_request = data.get('request')
    if not user_request:
        return jsonify({'error': 'No request provided'}), 400

    category = predict_category(user_request)
    return jsonify({'category': category})

if __name__ == '__main__':
    app.run(debug=True)

def send_message(chat_id, text, disable_notification=True):
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    data = {
        "chat_id": chat_id,
        "text": text,
        "disable_notification": disable_notification
    }
    response = requests.post(url, json=data)
    return response.json()

@app.route("/new_report", methods=["POST"])

def new_report():
    data = request.get_json()
    msg=f"""Сообщение пользователя: {data["text"]}
Форма обратной связи: {data["callback"]}"""
    send_message(1070122283, msg)
    return {"status": "ok"}

if __name__ == "__main__":
    app.run(debug=True,port=5000)
